from openpyxl.styles import PatternFill, Font , Alignment
from openpyxl.worksheet.worksheet import Worksheet

# Funciones para formatear Excel

# Aplicar formato al encabezado
def Aplicar_formato_encabezado(HojaActual : Worksheet):
    '''
    Función que aplica formato al encabezado de la hoja
    '''
            
    # Darle formato a los Títulos de las columnas
    Fondotitulo = PatternFill(start_color='002060' , end_color='002060' ,  fill_type='solid')
    LetraColor = Font(color='FFFFFF')

    for cell in HojaActual[1]:
        cell.fill = Fondotitulo
        cell.font = LetraColor


# Aplica formato de moneda a las columnas de importes
def Aplicar_formato_moneda(HojaActual : Worksheet , ColumnaInicial : int , ColumnaFinal : int):
    '''
    Función que aplica formato de moneda a las columnas de importes
    '''
    
    formato = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

    for cell in HojaActual.iter_rows(min_row=2, min_col=ColumnaInicial, max_row=HojaActual.max_row, max_col=ColumnaFinal):
        for celda in cell:
            celda.number_format = formato


# Autoajustar los anchos de las columnas según el contenido
def Autoajustar_columnas(HojaActual : Worksheet):
    '''
    Función que autoajusta las columnas de la hoja
    '''
    
    for column_cells in HojaActual.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        HojaActual.column_dimensions[column_cells[0].column_letter].width = length + 2


# Agregar filtros de datos a las hojas
def Agregar_filtros(HojaActual : Worksheet):
    '''
    Función que agrega filtros a la hoja
    '''
    
    HojaActual.auto_filter.ref = HojaActual.dimensions

# Alinear columnas
def Alinear_columnas(HojaActual : Worksheet , ColumnaInicial : int , ColumnaFinal : int , Alineacion : str ):
    '''
    Función que alinea las columnas de la hoja
    '''
    Alineacion = Alignment(horizontal=Alineacion)
    
    for cell in HojaActual.iter_rows(min_row=2, min_col=ColumnaInicial, max_row=HojaActual.max_row, max_col=ColumnaFinal):
        for celda in cell:
            celda.alignment = Alineacion